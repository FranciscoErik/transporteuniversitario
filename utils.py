import os
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re
from io import BytesIO
from openpyxl.utils import get_column_letter

def generate_pdf_report(operations):
    """Gera um relatório em PDF com os dados das operações."""
    # Criar buffer para armazenar o PDF
    buffer = BytesIO()
    
    # Preparar dados para o relatório
    data = [['Data', 'Ônibus', 'Custo Combustível', 'Salário Motorista', 'Salário Monitor', 'Custo Total']]
    
    for op in operations:
        data.append([
            op['date'],
            op['bus_plate'],
            format_currency(op['fuel_cost']),
            format_currency(op['driver_salary']),
            format_currency(op['monitor_salary']),
            format_currency(op['total_cost'])
        ])
    
    # Criar documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    table = Table(data)
    
    # Estilizar tabela
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)
    
    # Construir PDF
    elements = []
    elements.append(table)
    doc.build(elements)
    
    # Mover o ponteiro para o início do buffer
    buffer.seek(0)
    return buffer

def generate_excel_report(operations):
    """Gera um relatório em Excel com os dados das operações."""
    # Criar buffer para armazenar o Excel
    buffer = BytesIO()
    
    # Criar workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Operações"
    
    # Definir cabeçalhos
    headers = ['Data', 'Ônibus', 'Custo Combustível', 'Salário Motorista', 'Salário Monitor', 'Custo Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Adicionar dados
    for row, op in enumerate(operations, 2):
        ws.cell(row=row, column=1, value=op['date'])
        ws.cell(row=row, column=2, value=op['bus_plate'])
        ws.cell(row=row, column=3, value=format_currency(op['fuel_cost']))
        ws.cell(row=row, column=4, value=format_currency(op['driver_salary']))
        ws.cell(row=row, column=5, value=format_currency(op['monitor_salary']))
        ws.cell(row=row, column=6, value=format_currency(op['total_cost']))
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Salvar no buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def format_currency(value):
    """Formata um valor numérico como moeda brasileira."""
    return f'R$ {value:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')

def validate_cnpj(cnpj):
    """Valida um CNPJ."""
    # Remove caracteres não numéricos
    cnpj = re.sub(r'[^0-9]', '', cnpj)
    
    if len(cnpj) != 14:
        return False
    
    # Verifica se todos os dígitos são iguais
    if cnpj == cnpj[0] * 14:
        return False
    
    # Validação do primeiro dígito verificador
    soma = 0
    peso = 5
    for i in range(12):
        soma += int(cnpj[i]) * peso
        peso = peso - 1 if peso > 2 else 9
    
    digito1 = 11 - (soma % 11)
    if digito1 > 9:
        digito1 = 0
    
    if int(cnpj[12]) != digito1:
        return False
    
    # Validação do segundo dígito verificador
    soma = 0
    peso = 6
    for i in range(13):
        soma += int(cnpj[i]) * peso
        peso = peso - 1 if peso > 2 else 9
    
    digito2 = 11 - (soma % 11)
    if digito2 > 9:
        digito2 = 0
    
    if int(cnpj[13]) != digito2:
        return False
    
    return True

def create_upload_folder():
    """Cria a pasta de uploads se não existir."""
    upload_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
    os.makedirs(upload_folder, exist_ok=True)
    return upload_folder

def generate_pdf_students_report(students):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Título
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30
    )
    elements.append(Paragraph("Relatório de Alunos", title_style))
    
    # Data de geração
    date_style = ParagraphStyle(
        'CustomDate',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray
    )
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style))
    elements.append(Spacer(1, 20))
    
    # Tabela de alunos
    data = [['Nome', 'Escola', 'Ponto de Parada', 'Rota', 'Ônibus']]
    for student in students:
        data.append([
            student['name'],
            student['school_name'],
            student['stop_name'],
            student['route_name'],
            student['bus_plate']
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_students_report(students):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"
    
    # Cabeçalho
    headers = ['Nome', 'Escola', 'Ponto de Parada', 'Rota', 'Ônibus']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dados
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student['name'])
        ws.cell(row=row, column=2, value=student['school_name'])
        ws.cell(row=row, column=3, value=student['stop_name'])
        ws.cell(row=row, column=4, value=student['route_name'])
        ws.cell(row=row, column=5, value=student['bus_plate'])
    
    # Estilização
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer 